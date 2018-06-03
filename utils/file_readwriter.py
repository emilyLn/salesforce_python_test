'''
Created on 2018年3月27日
'''
import openpyxl
import yaml
import os
from utils.config import Config
from utils.log import logger

class SheetTypeError(Exception):
    pass
class SheetNumberError(Exception):
    pass

#打开excel文件，读文件内容，并回写文件中的部分内容
class ExcelReadWriter:
    '''
    使用openpyxl读写excel文档
    如：
    excel中内容为：
    | A  | B  | C  |
    | A1 | B1 | C1 |
    | A2 | B2 | C2 |

    如果 print(ExcelReadWriter(excel, title_line=True).data)，输出结果：
    [{A: A1, B: B1, C:C1}, {A:A2, B:B2, C:C2}]
    print(ExcelReadWriter(excel, title_line=True).title), 输出结果：
    [A, B, C]

    如果 print(ExcelReadWriter(excel, title_line=False).data)，输出结果：
    [[A,B,C], [A1,B1,C1], [A2,B2,C2]]
    print(ExcelReadWriter(excel, title_line=False).title), 输出结果：
    None

    可以指定sheet，通过index或者name：
    ExcelReadWriter(excel, sheet=2)
    ExcelReadWriter(excel, sheet='BaiDuTest')
    '''

    def __init__(self, excelpath, sheet=1, title_line=True):
        '''
        sheet = 1, 默认定位到第一个sheet页的数据
        title_line = True, 默认数据中包含表头，第一行为表头
        '''
        if os.path.exists(excelpath):
            self._excelpath = excelpath  # excel文件路径
        else:
            raise FileNotFoundError('文件不存在！')
        self._sheet = sheet  # sheet可以是int表示表格的索引，可以是str表示表格的名称
        self._title_line = title_line  # 是否存在标题行，有标题行，每一行都是都是对应列名的取值；没有标题行，每一行都是一个列表
        self._data = list()   # 用于存储每行生成的数据。
        self._title = {}  # 用于存储标题行的数据。
        
    @property
    def data(self):
        if not self._data: # 创建实例后，第一次调用时取数据，以后调用该实例，则直接从列表中取出数据即可
            self.fetchData()
        return self._data    
    @property
    def title(self):
        return self._title
    
    def fetchData(self):
        # load workbook
        workbook = openpyxl.load_workbook(self._excelpath)
        # load sheet
        if not isinstance(self._sheet, str) and not isinstance(self._sheet, int) :
            raise SheetTypeError('Please pass in <type int> or <type str>, not {0}'.format(type(self.sheet)))
        if isinstance(self._sheet, str):
            sheet = workbook.get_sheet_by_name(self._sheet)  # get_sheet_by_name直接通过工作表名字取
        else:
            sheet_num = len(workbook.sheetnames)
            if self._sheet > sheet_num or self._sheet < 1:
                raise SheetNumberError('The sheet number you want to load is less than 0 or larger than total number in the file')
            sheet = workbook[workbook.sheetnames[self._sheet - 1]] # Openpyxl模块支持类似字典键值对映射的方式，来获取表格的内容
        
        # fetch title and data
        if self._title_line: # 包含表头行
            for cellObj in tuple(sheet.rows)[0]:
                if cellObj.value in self._title:
                    self._title.get(cellObj.value).append(cellObj.coordinate) # cell.column返回列名（例：B） cell.row返回行号（例：4） cell.coordiante()返回单元格的定位（例：B4)
                else:
                    self._title.setdefault(cellObj.value,[]).append(cellObj.coordinate)
            
            for row in sheet.iter_rows(min_row=2):
                rowValues=[]
                for cellObj in row:
                    rowValues.append(cellObj.value)
                self._data.append(dict(zip(self._title.keys(), rowValues)))
        else:
            self._title = {}
            for row in tuple(sheet.rows):
                rowValues = []
                for cellObj in row:
                    rowValues.append(cellObj.value)
                
                self._data.append(rowValues)
        return self._data
    
if __name__ == '__main__':
    e = 'C:/eclipse/workspace/salesforce_python_test/data/test_data/test_data_dailyworkplan.xlsx'
    reader = ExcelReadWriter(e, title_line=False)
    print(reader.data)
    print(reader.title)