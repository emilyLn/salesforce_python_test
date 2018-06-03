# coding:UTF-8
from selenium import webdriver
from .gloVar import gloVar
import openpyxl
import time

class Common:
    global driver
    global currentWindowHandle
    def __init__(self):
        self.driver = webdriver.Chrome()
        driver = self.driver
        currentWindowHandle = self.driver.current_window_handle
        self.Excel = Excel()
        # self.TXT = TXT()
        self.LoginPage = LoginPage()
        self.Menu = Menu()
        self.DailyWorkPlanPage = DailyWorkPlanPage()
        self.CreatePlanPage = CreatePlanPage()
        self.DailyWorkPlanDetailInformation = DailyWorkPlanDetailInformation()

    def get(self,strUrl):
        self.driver.get(strUrl)
        print("driver.get(" + strUrl + ")")

    def timeOut(self,sec):
        self.driver.implicitly_wait(sec)
        print ("driver.implicitly_wait(" + str(sec) + ")")

    def maximizeWindow(self):
        self.driver.maximize_window()
        print ("driver.maximize_window()")

    def switchWindow(self):
        handles = self.driver.window_handles
        for handle in handles:
            if handle != currentWindowHandle: 
                self.driver.switch_to.window(handle)
                self.currentWindow = False
        print ("driver.swtichWindow()")

    def getCurrentURL(self):
        result = self.driver.current_url
        print ("driver.getCurrentURL : " + result)
        return result

    def closeWindow(self):
        gloVar.currentWindowHandle = self.driver.current_window_handle
        self.driver.close()
        print ("driver.closWindow()")

    def sleep(self,sec):
        time.sleep(sec)
        print ("time.sleep(" + str(sec) + ")")
    
class Element:
    global driver
    def __init__(self,locator,name):
        # gloVar.driver.switch_to.frame("top")
        self.ele = driver.find_element_by_xpath(locator)
        self.name = name
        self.locator = locator

    def sendKeys(self,txt):
        self.ele.send_keys(txt)
        print (self.name + ".send_keys(" + txt + ")")

    def click(self):
        self.ele.click()
        print (self.name + ".click()")

    def select(self, keyword):
        global driver
        self.ele.click()
        optionEle = driver.find_element_by_xpath(self.locator + "/option[contains(text(),'" + keyword + "')]")
        optionEle.click()
        print (self.name + ".select(" + keyword + ")")

    def getText(self):
        result = self.ele.text
        print (self.name + ".getText() :")
        print (result)
        return (result)

class LoginPage:
    def __init__(self):
        self.name = "LoginPage"
    def userId(self):
        return Element("//input[@id='username']",self.name + "." + "userId")
    def password(self):
        return Element("//input[@id='password']",self.name + "." + "password")
    def loginButton(self):
        return Element("//input[@id='Login']",self.name + "." + "Login")

class Menu:
    def __init__(self):
        self.name = "Menu"
    def MoreTabs_Tab(self):
        return Tab("//li[@id='MoreTabs_Tab']/a",self.name + "." + "MoreTabs_Tab")

# Menu > Tab
class Tab(Element):
    global driver
    def __init__(self,locator,name):
        self.locator = locator + "/.."
        self.name = name
        self.ele = driver.find_element_by_xpath(locator)
    def Item(self, keyword):
        return Element(self.locator + "//a[text()='" + keyword + "']", self.name + "." + "Item(" + keyword + ")")

class DailyWorkPlanPage:
    def __init__(self):
        self.name = "DailyWorkPlanPage"
    def newButton(self):
        return Element("//input[contains(@value,'新建')]",self.name + "." + "newButton")

class CreatePlanPage:
    def __init__(self):
        self.name = "CreatePlanPage"
    def saveButton(self):
        return Element("//input[@name='page:form:pageblock:j_id6:bottom:j_id7']",self.name + "." + "saveButton")
    def startDate(self):
        return Element("//input[@id='page:form:pageblock:PlanInfo:j_id9:Plan_Date__c']",self.name + "." + "startDate")
    def startTime(self):
        return Element("//select[@id='page:form:pageblock:PlanInfo:j_id13:Plan_Start_Time__c']",self.name + "." + "startTime")
    def endTime(self):
        return Element("//select[@id='page:form:pageblock:PlanInfo:j_id16:Plan_End_Time__c']",self.name + "." + "endTime")
    def workContent(self):
        return Element("//select[@id='page:form:pageblock:PlanInfo:j_id19:Plan_Content__c']",self.name + "." + "workContent")
    def table(self):
        return Element("//table[contains(@class,'detailList')]",self.name + "." + "table")

class DailyWorkPlanDetailInformation:
    def __init__(self):
        self.name = "DailyWorkPlanDetailInformation"
    def table(self,keyword):
        return DailyWorkPlanDetailInformationTable("//h3[text()='" + keyword  + "']/ancestor::div[contains(@class,'brandTertiaryBrd')]/following::div[1]",self.name + "." + "table(" + keyword + ")")

class DailyWorkPlanDetailInformationTable(Element):
    def __init__(self,locator,name):
        global driver
        self.name = "Table"
        self.ele = driver.find_element_by_xpath(locator)
        self.name = name
        self.locator = locator
    def field(self,keyword):
        return Element(self.locator + "//td[text()='" + keyword + "']/following::td[1]",self.name + "." + "field(" + keyword + ")")

class TXT:
    def __init__(self,path):
        self.path = path

    def read(self):
        f = open(self.path)
        result = f.read()
        f.close
        return result

    def write(self,msg):
        f = open(self.path,"w")
        f.write(msg)
        f.close
        print ("TXT.write() : " + msg)

    def add(self,msg):
        f = open(self.path)
        orgMsg = f.read()
        f.close
        f = open(self.path,"w")
        f.write(orgMsg + "\n" + msg)
        f.close
        print ("TXT.add() : " + msg)
    

class Excel:
    def __init__(self):
        self.name = "Excel"

    def workbook(self,path):
        return Workbook(path)

# Excel.Workbook
class Workbook:
    def __init__(self,path):
        self.wb = openpyxl.load_workbook(filename=path)
        self.path = path

    def worksheet(self,shtName):
        return Worksheet(shtName,self.wb,self.path)

# Excel.Workbook.Worksheet
class Worksheet:
    def __init__(self,shtName,wb,path):
        self.__sht = wb.get_sheet_by_name(shtName)
        self.__wb = wb
        self.__path = path
        self.name = self.__sht.title
        self.lastRow = self.__sht.max_row
        self.lastCol = self.__sht.max_column
    def cell(self,rowIndex, colIndex):
        return Cell(rowIndex, colIndex, self.__wb, self.__sht, self.__path)

# Excel.Workbook.Worksheet.Cell
class Cell:
    def __init__(self, rowIndex, colIndex, wb, sht, path):
        self.rowIndex = rowIndex
        self.colIndex = colIndex
        self.__wb = wb
        self.__sht = sht
        self.__path = path

    def getValue(self): 
        result = self.__sht.cell(row=self.rowIndex, column=self.colIndex).value
        return result

    def setValue(self, txt):
        self.__sht.cell(row=self.rowIndex,column=self.colIndex, value=txt)
        self.__wb.save(self.__path)